import random as r # Случайный цвет для класса объектов
from collections import defaultdict, deque # Фиксированный массив для трекинга
from time import time # Расчет кадров в секунду
from ultralytics import YOLO, checks # Нейросеть YOLO
import cv2 # Чтение видеопотока и визуализация результатов YOLO
import numpy as np # Преобразования матриц
from tqdm import tqdm # ProgressBar
from openpyxl import Workbook # Запись результатов


# Функция для визуализации направления движения объекта
def draw_direction(frame, trail, color):
    movements = [] # Инициализация списка для векторов перемещения
    trail = list(trail) # Получение массива центров объекта

    # Вычисление векторов перемещения по последним 10 точкам
    for i in range(len(trail)-10, len(trail)):
        movements.append(np.array(trail[i]) - np.array(trail[i - 1]))

    movement = np.mean(movements, axis=0) # Усреднение векторов перемещения

    # Фильтруем тряску (Не показываем направление если движение незначительное)
    if abs(movement).astype(float)[0] > 0.3 or abs(movement).astype(float)[1] > 0.3:
        # Определение координат конца стрелки
        point = tuple((np.array(trail[-1]) + 10 * movement).astype(int))

        # Нарисовать стрелку, указывающую на направление движения
        cv2.arrowedLine(img=frame, pt1=trail[-1], pt2=point, color=color, thickness=2)

# Функция для визуализации траектории объекта
def draw_trace(frame, trail, color):
    # Конвертируем массив точек для функции cv2.polylines()
    points = np.hstack(trail).astype(np.int32).reshape((-1, 1, 2))

    # Рисуем траекторию
    cv2.polylines(img=frame, pts=[points], isClosed=False, color=color, thickness=2)

# Функция для визуализации ограничивающей рамки объекта
def draw_bbox(frame, id, name, conf, color, x0, y0, x1, y1):
    # Подпись с уникальным номером, классом объекта и уверенностью в процентах
    text = f'{name} #{id} - {int(conf*100)}%'
    
    # Ограничивающая рамка
    cv2.rectangle(img=frame, pt1=(x0, y0), pt2=(x1, y1), color=color, thickness=1)
    
    # Плашка под текст
    frame[(y0 - 20):(y0), 
          (x0):(x0 + cv2.getTextSize(text, cv2.FONT_ITALIC, 0.5, 1)[0][0])] = color
                
    # Добавляем текст
    cv2.putText(img=frame, text=text, org=(x0, y0-5), 
                fontFace=cv2.FONT_ITALIC, fontScale=0.5, color=0, thickness=1)

# Функция для визуализации центра объекта
def draw_centr(frame, color, cx, cy):
    text = f'{cx, cy}' # Подпись с координатой x, y центра объекта
    
    # Плашка под текст
    frame[(cy - 30):(cy - 10),
          (cx + 10):(cx + cv2.getTextSize(text, cv2.FONT_ITALIC, 0.5, 1)[0][0] + 10)] = color

    # Добавляем текст
    cv2.putText(img=frame,text=text,org=(cx + 10, cy - 15),
                fontFace=cv2.FONT_ITALIC, fontScale=0.5, color=0, thickness=1)
    
    cv2.circle(frame, (cx, cy), 4, color, -1) # Рисуем точку центра объекта

# Создаем новую книгу Excel
wb = Workbook()

# Выбираем активный лист
ws = wb.active

# Добавляем заголовки столбцов
ws.append(["Time", "Obj", "MaxID", "aFPS"])

# Состояния по умолчанию
show_trace = True
show_centr = True
show_bbox = True
show_direction = True

checks() # Проверяем используется ли CUDA
name_model = 'yolov8m'
tracker = 'bytetrack'
model = YOLO(f'models\\{name_model}.pt') # Подгружаем модель YOLOv8
model.fuse() # Объединение слоев для оптимизации вывода.

file = 'cam1(original size).mp4'
# Открываем видеопоток и создаем объект для сохранения
cap = cv2.VideoCapture(f'input\\{file}')
cap.set(cv2.CAP_PROP_POS_FRAMES, 675) # Выбор кадра для старта, если надо с начала - закоментить.
frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

# Буфер точек положения центра объекта
trails = defaultdict(lambda: deque(maxlen=50))

# Переменные для подсчета кадров в секунду
global_start_time = time()
best_obj = 0
iterat = 2
# Основной цикл программы (15 секунд видео, если надо всё - заменить на "range(frames-1)")
for i in tqdm(range(375), unit=' frames'):
    ret, frame = cap.read()
    if not ret:break

    obj = 0 # Обнуляем количество объектов

    # Детекция и трекинг объекта средствами YOLO и (botsort/bytetrack)
    results = model.track(source=frame, iou=0.6, conf=0.6, 
                          persist=True, verbose=False, tracker=f'config\\{tracker}.yaml')

    # Вытаскиваем ограничивающие рамки, классы, уникальные номера и уверенности
    if results[0].boxes.id is not None:
        boxes = results[0].boxes.xywh.cpu().numpy().astype(int)
        cls = results[0].boxes.cls.cpu().numpy().astype(int)
        ids = results[0].boxes.id.cpu().numpy().astype(int)
        confs = results[0].boxes.conf.cpu().numpy().astype(float)
        names = results[0].names
        obj = len(ids)

        # Обрабатываем информацию каждого объекта
        for box, cl, id, conf in zip(boxes, cls, ids, confs):
            name = names[cl] # Имя класса объекта
            cx,cy,w,h = box # Координата центра, ширина и высота ограничивающей рамки
            
            # Крайние точки для ограничивающей рамки
            x0, y0, = int(cx - w/2), int(cy - h/2)
            x1, y1 = int(cx + w/2), int(cy + h/2)

            # Задаём цвета для красоты (Единый цвет у класса)
            r.seed(int(cl))
            color = (r.randint(150, 255), r.randint(150, 255), r.randint(150, 255))

            trails[id].append((cx, cy)) # Добавляем координату центра объекта в массив
            
            # Отображение направления движения
            if show_direction and len(list(trails[id]))>10:draw_direction(frame, trails[id], color)
            
            # Отображение траектории
            if show_trace:draw_trace(frame, trails[id], color)
                            
            # Отображение ограничивающей рамки
            if show_bbox:draw_bbox(frame, id, name, conf, color, x0, y0, x1, y1)
                
            # Отображение центра объекта с координатой
            if show_centr:draw_centr(frame, color, cx, cy)

    if obj > best_obj: best_obj = obj
    
    if i % 25 == 0:
        ws[f'A{iterat}'] = iterat-2
        ws[f'B{iterat}'] = best_obj
        ws[f'C{iterat}'] = len(trails)
        
        iterat += 1
        best_obj=0

ws[f'A{iterat}'] = iterat-2
ws[f'B{iterat}'] = best_obj
ws[f'C{iterat}'] = len(trails)
ws['D2'] = (376 / (time() - global_start_time))
cap.release() # Высвобождаем ресурсы

output_path = f'output\\{file.split(".")[0]}_{name_model}_{tracker}.xlsx'

# Сохраняем книгу
wb.save(output_path)

print('Тест завершен!')
cv2.destroyAllWindows() # Закрываем все окна